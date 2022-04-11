# 爬取Arduino库的信息并分类整理到Excel文件中
# 附带爬取这些库在github中的star数量
import requests
import openpyxl

def set_add_with_check(s, x):
  l = len(s)
  s.add(x)
  return len(s) != l

def get_github_star(git_url):
    if 'github' not in git_url:
        return None

    headers = {
        'Authorization':'token '+'your TokenID',
        'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'
    }

    user_repo = git_url[len('https://github.com/'):] # user/repo
    repo_info_api = 'https://api.github.com/repos/' + user_repo #拼出获取仓库信息的API
    repo_info_api = repo_info_api.rstrip('\r\n') # 去除行末尾的\r\n或\n，否则会导致url错误
    repo_info_api = repo_info_api.rstrip('\n')
    repo_info_api = repo_info_api.rstrip('.git') # 去除行末尾可能潜在的.git
    repo_info_dict = requests.get(repo_info_api, headers = headers).json()

    try:
        stargazers_count = repo_info_dict['stargazers_count'] # 获取到repo的star数量
    except:
        return None
    return stargazers_count

wb = openpyxl.Workbook()

sheet_display = wb.create_sheet('Display', 0)
sheet_communication = wb.create_sheet('Communication', 1)
sheet_signalio = wb.create_sheet('Signal Input-Output', 2)
sheet_sensors = wb.create_sheet('Sensors', 3)
sheet_devicecontrol = wb.create_sheet('Device Control', 4)
sheet_timing = wb.create_sheet('Timing', 5)
sheet_datastorage = wb.create_sheet('Data Storage', 6)
sheet_dataprocessing = wb.create_sheet('Data Processing', 7)
sheet_others = wb.create_sheet('Others', 7)
ws = wb['Sheet']
wb.remove(ws)

wb.save('index.xlsx')

# 从Arduino官网获取最新库信息
arduino_dict = requests.get("http://downloads.arduino.cc/libraries/library_index.json").json()
arduino_attributes_list = arduino_dict['libraries']

repo_category_display = set()
repo_category_communication = set()
repo_category_signalIO = set()
repo_category_sensors = set()
repo_category_devicecontrol = set()
repo_category_timing = set()
repo_category_datastorage = set()
repo_category_dataprocessing = set()
repo_category_other = set()

sheet_display_row_counter = 1
sheet_communication_row_counter = 1
sheet_signalIO_row_counter = 1
sheet_sensors_row_counter = 1
sheet_devicecontrol_row_counter = 1
sheet_timing_row_counter = 1
sheet_datastorage_row_counter = 1
sheet_dataprocessing_row_counter = 1
sheet_other_row_counter = 1

for arduino_library_dist in arduino_attributes_list:
    architecture = arduino_library_dist['architectures']

    # 专用于某个Arduino特定的板子，无法通用，跳过
    if len(architecture) != 0 and architecture[0] != '*':
        continue

    # 按照category的分类，放置于不同的set中
    repo_git = arduino_library_dist['repository']
    if arduino_library_dist['category'] == 'Display':
        if set_add_with_check(repo_category_display, repo_git):
            sheet_display.cell(row=sheet_display_row_counter, column=1).value = arduino_library_dist['name']
            sheet_display.cell(row=sheet_display_row_counter, column=2).value = repo_git
            sheet_display.cell(row=sheet_display_row_counter, column=3).value = arduino_library_dist['sentence']
            try:
                sheet_display.cell(row=sheet_display_row_counter, column=4).value = arduino_library_dist['paragraph']
            except:
                sheet_display.cell(row=sheet_display_row_counter, column=4).value = None
            sheet_display.cell(row=sheet_display_row_counter, column=5).value = arduino_library_dist['author']
            sheet_display.cell(row=sheet_display_row_counter, column=6).value = get_github_star(repo_git)
            sheet_display_row_counter += 1

    elif arduino_library_dist['category'] == 'Communication':
        if set_add_with_check(repo_category_communication, repo_git):
            sheet_communication.cell(row=sheet_communication_row_counter, column=1).value = arduino_library_dist['name']
            sheet_communication.cell(row=sheet_communication_row_counter, column=2).value = repo_git
            sheet_communication.cell(row=sheet_communication_row_counter, column=3).value = arduino_library_dist['sentence']
            try:
                sheet_communication.cell(row=sheet_communication_row_counter, column=4).value = arduino_library_dist['paragraph']
            except:
                sheet_communication.cell(row=sheet_communication_row_counter, column=4).value = None
            sheet_communication.cell(row=sheet_communication_row_counter, column=5).value = arduino_library_dist['author']
            sheet_communication.cell(row=sheet_communication_row_counter, column=6).value = get_github_star(repo_git)
            sheet_communication_row_counter += 1

    elif arduino_library_dist['category'] == 'Signal Input/Output':
        if set_add_with_check(repo_category_signalIO, repo_git):
            sheet_signalio.cell(row=sheet_signalIO_row_counter, column=1).value = arduino_library_dist['name']
            sheet_signalio.cell(row=sheet_signalIO_row_counter, column=2).value = repo_git
            sheet_signalio.cell(row=sheet_signalIO_row_counter, column=3).value = arduino_library_dist['sentence']
            try:
                sheet_signalio.cell(row=sheet_signalIO_row_counter, column=4).value = arduino_library_dist['paragraph']
            except:
                sheet_signalio.cell(row=sheet_signalIO_row_counter, column=4).value = None
            sheet_signalio.cell(row=sheet_signalIO_row_counter, column=5).value = arduino_library_dist['author']
            sheet_signalio.cell(row=sheet_signalIO_row_counter, column=6).value = get_github_star(repo_git)
            sheet_signalIO_row_counter += 1

    elif arduino_library_dist['category'] == 'Sensors':
        if set_add_with_check(repo_category_sensors, repo_git):
            sheet_sensors.cell(row=sheet_sensors_row_counter, column=1).value = arduino_library_dist['name']
            sheet_sensors.cell(row=sheet_sensors_row_counter, column=2).value = repo_git
            sheet_sensors.cell(row=sheet_sensors_row_counter, column=3).value = arduino_library_dist['sentence']
            try:
                sheet_sensors.cell(row=sheet_sensors_row_counter, column=4).value = arduino_library_dist['paragraph']
            except:
                sheet_sensors.cell(row=sheet_sensors_row_counter, column=4).value = None
            sheet_sensors.cell(row=sheet_sensors_row_counter, column=5).value = arduino_library_dist['author']
            sheet_sensors.cell(row=sheet_sensors_row_counter, column=6).value = get_github_star(repo_git)
            sheet_sensors_row_counter += 1

    elif arduino_library_dist['category'] == 'Device Control':
        if set_add_with_check(repo_category_devicecontrol, repo_git):
            sheet_devicecontrol.cell(row=sheet_devicecontrol_row_counter, column=1).value = arduino_library_dist['name']
            sheet_devicecontrol.cell(row=sheet_devicecontrol_row_counter, column=2).value = repo_git
            sheet_devicecontrol.cell(row=sheet_devicecontrol_row_counter, column=3).value = arduino_library_dist['sentence']
            try:
                sheet_devicecontrol.cell(row=sheet_devicecontrol_row_counter, column=4).value = arduino_library_dist['paragraph']
            except:
                sheet_devicecontrol.cell(row=sheet_devicecontrol_row_counter, column=4).value = None
            sheet_devicecontrol.cell(row=sheet_devicecontrol_row_counter, column=5).value = arduino_library_dist['author']
            sheet_devicecontrol.cell(row=sheet_devicecontrol_row_counter, column=6).value = get_github_star(repo_git)
            sheet_devicecontrol_row_counter += 1

    elif arduino_library_dist['category'] == 'Timing':
        if set_add_with_check(repo_category_timing, repo_git):
            sheet_timing.cell(row=sheet_timing_row_counter, column=1).value = arduino_library_dist['name']
            sheet_timing.cell(row=sheet_timing_row_counter, column=2).value = repo_git
            sheet_timing.cell(row=sheet_timing_row_counter, column=3).value = arduino_library_dist['sentence']
            try:
                sheet_timing.cell(row=sheet_timing_row_counter, column=4).value = arduino_library_dist['paragraph']
            except:
                sheet_timing.cell(row=sheet_timing_row_counter, column=4).value = None
            sheet_timing.cell(row=sheet_timing_row_counter, column=5).value = arduino_library_dist['author']
            sheet_timing.cell(row=sheet_timing_row_counter, column=6).value = get_github_star(repo_git)
            sheet_timing_row_counter += 1

    elif arduino_library_dist['category'] == 'Data Storage':
        if set_add_with_check(repo_category_datastorage, repo_git):
            sheet_datastorage.cell(row=sheet_datastorage_row_counter, column=1).value = arduino_library_dist['name']
            sheet_datastorage.cell(row=sheet_datastorage_row_counter, column=2).value = repo_git
            sheet_datastorage.cell(row=sheet_datastorage_row_counter, column=3).value = arduino_library_dist['sentence']
            try:
                sheet_datastorage.cell(row=sheet_datastorage_row_counter, column=4).value = arduino_library_dist['paragraph']
            except:
                sheet_datastorage.cell(row=sheet_datastorage_row_counter, column=4).value = None
            sheet_datastorage.cell(row=sheet_datastorage_row_counter, column=5).value = arduino_library_dist['author']
            sheet_datastorage.cell(row=sheet_datastorage_row_counter, column=6).value = get_github_star(repo_git)
            sheet_datastorage_row_counter += 1

    elif arduino_library_dist['category'] == 'Data Processing':
        if set_add_with_check(repo_category_dataprocessing, repo_git):
            sheet_dataprocessing.cell(row=sheet_dataprocessing_row_counter, column=1).value = arduino_library_dist['name']
            sheet_dataprocessing.cell(row=sheet_dataprocessing_row_counter, column=2).value = repo_git
            sheet_dataprocessing.cell(row=sheet_dataprocessing_row_counter, column=3).value = arduino_library_dist['sentence']
            try:
                sheet_dataprocessing.cell(row=sheet_dataprocessing_row_counter, column=4).value = arduino_library_dist['paragraph']
            except:
                sheet_dataprocessing.cell(row=sheet_dataprocessing_row_counter, column=4).value = None
            sheet_dataprocessing.cell(row=sheet_dataprocessing_row_counter, column=5).value = arduino_library_dist['author']
            sheet_dataprocessing.cell(row=sheet_dataprocessing_row_counter, column=6).value = get_github_star(repo_git)
            sheet_dataprocessing_row_counter += 1

    else:
        if set_add_with_check(repo_category_other, repo_git):
            sheet_others.cell(row=sheet_other_row_counter, column=1).value = arduino_library_dist['name']
            sheet_others.cell(row=sheet_other_row_counter, column=2).value = repo_git
            sheet_others.cell(row=sheet_other_row_counter, column=3).value = arduino_library_dist['sentence']
            try:
                sheet_others.cell(row=sheet_other_row_counter, column=4).value = arduino_library_dist['paragraph']
            except:
                sheet_others.cell(row=sheet_other_row_counter, column=4).value = None
            sheet_others.cell(row=sheet_other_row_counter, column=5).value = arduino_library_dist['author']
            sheet_others.cell(row=sheet_other_row_counter, column=6).value = get_github_star(repo_git)
            sheet_other_row_counter += 1

wb.save('repoindex.xlsx')
wb.close()
